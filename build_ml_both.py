import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook
import os

os.chdir(r'd:\cc_test')

# ---------- 读取 ml_catalyst_yield.xlsx ----------
wb_src = openpyxl.load_workbook('ml_catalyst_yield.xlsx', read_only=True, data_only=True)
ws_src = wb_src['ML_data']
all_rows = list(ws_src.iter_rows(values_only=True))
header = list(all_rows[0])
data = [list(r) for r in all_rows[1:]]
n = len(data)

# 计算各列缺失率
missing_pct = []
for j in range(len(header)):
    cnt = sum(1 for row in data if row[j] is None)
    missing_pct.append(cnt / n * 100)

# 目标列（C 开头的列）
target_cols = [j for j, h in enumerate(header) if h.startswith('C')]
# 特征列
feature_cols = [j for j, h in enumerate(header) if not h.startswith('C')]
# 配方类特征列（M/FM/Z/FZ）
comp_cols = [j for j, h in enumerate(header)
             if h.split('_')[0] in
             ['M1-1','M1-2','M1-3','M2','M3-1','M3-2','M4','M5',
              'FM1','FM2','FM3','FM4','FM5','FM6','FM7','FM8','FM9','FM10','FM11',
              'Z1','Z2','Z3','Z4','Z5','Z6','Z7','Z8','Z9','Z10','Z11','Z12','Z13',
              'FZ1','FZ2','FZ3','FZ4','FZ5','FZ6','FZ7','FZ8']]


def write_info_sheet(ws, col_indices, header, missing_pct):
    ws.append(['列名', '类型', '缺失率(%)', '说明'])
    for j in col_indices:
        h = header[j]
        code = h.split('_')[0]
        if h.startswith('C'):
            t = '目标变量'
            note = '完整，直接用于训练'
        elif code in ['T1','T2']:
            t = '特征-操作条件'
            note = '完整'
        elif code == 'oil':
            t = '特征-原料油(类别)'
            note = '建议 LabelEncoder 或 get_dummies 编码'
        elif j in comp_cols:
            t = '特征-配方组成'
            note = 'NaN已填0（未添加该组分）'
        else:
            t = '特征-理化性质'
            note = 'NaN=未测量，XGBoost可原生处理'
        ws.append([h, t, f'{missing_pct[j]:.1f}', note])


# ============================================================
# 方案 A：全部 768 行，配方列 NaN 填 0，删 >95% 缺失列
# ============================================================
drop_set = set(j for j in feature_cols if missing_pct[j] > 95)
keep_A = [j for j in range(len(header)) if j not in drop_set]

wb_A = Workbook()
ws_A = wb_A.active
ws_A.title = 'ML_data'
ws_A.append([header[j] for j in keep_A])

written_A = 0
for row in data:
    out = []
    for j in keep_A:
        val = row[j]
        # 配方列 NaN -> 0
        if val is None and j in comp_cols:
            val = 0
        out.append(val)
    ws_A.append(out)
    written_A += 1

ws_A_info = wb_A.create_sheet('Column_Info')
write_info_sheet(ws_A_info, keep_A, header, missing_pct)

ws_A_note = wb_A.create_sheet('说明')
notes = [
    ['方案A：全量数据（768行）'],
    [''],
    ['特点', '说明'],
    ['行数', '768'],
    ['特征列数', str(len([j for j in keep_A if not header[j].startswith('C')]))],
    ['目标列数', str(len(target_cols))],
    ['配方列处理', 'NaN填0（未添加该组分=0，符合催化剂配方逻辑）'],
    ['理化性质列', '保留NaN，用XGBoost/LightGBM训练时无需额外处理'],
    ['删除的列', '缺失率>95%的列（信息量极少，影响模型质量）'],
    ['推荐模型', 'XGBoost / LightGBM（原生支持NaN和稀疏特征）'],
    [''],
    ['推荐目标变量（单独预测）', ''],
    ['C4_汽油', '汽油产率，最重要'],
    ['C9_转化率', '总转化率'],
    ['C3_液化气', '液化气产率'],
    ['C1_焦炭', '焦炭产率（希望越低越好）'],
]
for row in notes:
    ws_A_note.append(row)

wb_A.save('ml_A_all.xlsx')
print(f'方案A: {written_A}行 x {len(keep_A)}列 -> ml_A_all.xlsx')


# ============================================================
# 方案 B：筛选"近完整"子集，可用任何模型
# 核心特征列：缺失率<50%的列（含所有目标列）
# ============================================================
core_feature_cols = [j for j in feature_cols
                     if missing_pct[j] < 50 or header[j] in ['T1_反应温度','T2_剂油比','oil_原料油']]
keep_B_cols = core_feature_cols + target_cols

# 筛选行：核心特征列（非target、非oil）中至少70%有值的行
check_cols = [j for j in core_feature_cols if header[j] not in ['T1_反应温度','T2_剂油比','oil_原料油']]

wb_B = Workbook()
ws_B = wb_B.active
ws_B.title = 'ML_data'
ws_B.append([header[j] for j in keep_B_cols])

written_B = 0
for row in data:
    # 要求 T1/T2 必须有值，且核心配方列至少1个有值
    t1_idx = header.index('T1_反应温度')
    t2_idx = header.index('T2_剂油比')
    if row[t1_idx] is None or row[t2_idx] is None:
        continue
    # 核心配方列有至少1个非空
    if check_cols and all(row[j] is None for j in check_cols):
        continue
    out = []
    for j in keep_B_cols:
        val = row[j]
        if val is None and j in comp_cols:
            val = 0
        out.append(val)
    ws_B.append(out)
    written_B += 1

ws_B_info = wb_B.create_sheet('Column_Info')
write_info_sheet(ws_B_info, keep_B_cols, header, missing_pct)

ws_B_note = wb_B.create_sheet('说明')
notes_B = [
    ['方案B：精简特征子集'],
    [''],
    ['特点', '说明'],
    ['行数', str(written_B)],
    ['特征列数', str(len(core_feature_cols))],
    ['目标列数', str(len(target_cols))],
    ['筛选逻辑', '只保留缺失率<50%的特征列；行要求T1/T2完整且至少一个配方特征有值'],
    ['配方列处理', 'NaN填0'],
    ['推荐模型', '随机森林 / 线性回归 / SVR / 神经网络（数据更干净，适合更多模型）'],
    [''],
    ['保留的特征列', '缺失率'],
]
for j in core_feature_cols:
    notes_B.append([header[j], f'{missing_pct[j]:.1f}%'])
for row in notes_B:
    ws_B_note.append(row)

wb_B.save('ml_B_精简版.xlsx')
print(f'方案B: {written_B}行 x {len(keep_B_cols)}列 -> ml_B_精简版.xlsx')
