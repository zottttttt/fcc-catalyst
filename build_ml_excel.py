import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook
import os

os.chdir(r'd:\cc_test')

wb_in = openpyxl.load_workbook('数据录入模板总-202509.xlsx', read_only=True, data_only=True)
ws_in = wb_in['Sheet1']

all_rows = list(ws_in.iter_rows(values_only=True))
data_rows = all_rows[3:]  # from row 4 (index 3)

# (col_index, code, name, type, fill_zero)
columns = [
    # 基质原料 (fill 0)
    (2,  'M1-1', 'S-1土',        'feature_composition', False),
    (3,  'M1-2', 'S-2土',        'feature_composition', False),
    (4,  'M1-3', '茂名土',       'feature_composition', False),
    (5,  'M2',   '埃洛石',       'feature_composition', False),
    (6,  'M3-1', '山西拟薄',     'feature_composition', False),
    (7,  'M3-2', '山东拟薄',     'feature_composition', False),
    (8,  'M4',   '铝溶胶',       'feature_composition', False),
    (9,  'M5',   '多孔',         'feature_composition', False),
    # 功能基质 (fill 0)
    (10, 'FM1',  '大孔氧化铝',    'feature_composition', False),
    (11, 'FM2',  'APM-7',        'feature_composition', False),
    (12, 'FM3',  'APM-9',        'feature_composition', False),
    (13, 'FM4',  '薄水',         'feature_composition', False),
    (14, 'FM5',  '大孔拟薄',     'feature_composition', False),
    (15, 'FM6',  '大比表硅铝',   'feature_composition', False),
    (16, 'FM7',  '基质稀土',     'feature_composition', False),
    (17, 'FM8',  '氨水沉淀稀土', 'feature_composition', False),
    (18, 'FM9',  '后沉淀稀土',   'feature_composition', False),
    (19, 'FM10', '白土细粉',     'feature_composition', False),
    (20, 'FM11', '草酸',         'feature_composition', False),
    # 分子筛 (fill 0)
    (22, 'Z1',  'RDSY(NK)',   'feature_composition', False),
    (23, 'Z2',  'RDSY',       'feature_composition', False),
    (24, 'Z3',  'MASY',       'feature_composition', False),
    (25, 'Z4',  'HRSY-1',     'feature_composition', False),
    (26, 'Z5',  'HRSY-3',     'feature_composition', False),
    (27, 'Z6',  'HRSY-4',     'feature_composition', False),
    (28, 'Z7',  'HRSY-5',     'feature_composition', False),
    (29, 'Z8',  'REY',        'feature_composition', False),
    (30, 'Z9',  'HASY-8',     'feature_composition', False),
    (31, 'Z10', 'GUSY-2',     'feature_composition', False),
    (32, 'Z11', 'GUSY-1',     'feature_composition', False),
    (33, 'Z12', '分子筛铝溶胶', 'feature_composition', False),
    (34, 'Z13', 'Beta分子筛',  'feature_composition', False),
    # 功能分子筛/改性剂 (fill 0)
    (35, 'FZ1', '介孔分子筛',   'feature_composition', False),
    (36, 'FZ2', '高硅ZSM-5',   'feature_composition', False),
    (37, 'FZ3', '低硅ZSM-5',   'feature_composition', False),
    (38, 'FZ4', 'SA-5',        'feature_composition', False),
    (39, 'FZ5', '磷铁改性择型', 'feature_composition', False),
    (40, 'FZ6', 'MgO',         'feature_composition', False),
    (41, 'FZ7', 'P2O5',        'feature_composition', False),
    (42, 'FZ8', '氧化钇',       'feature_composition', False),
    # 理化性质 (keep NaN)
    (43, 'L1',  'Na',     'feature_property', False),
    (44, 'L2',  'P',      'feature_property', False),
    (45, 'L3',  'RE',     'feature_property', False),
    (49, 'L7',  '强度',   'feature_property', False),
    (50, 'L8',  '比表',   'feature_property', False),
    (53, 'L11', '孔体积', 'feature_property', False),
    (55, 'L13', '微活',   'feature_property', False),
    # 操作条件
    (56, 'T1',  '反应温度', 'feature_condition', False),
    (57, 'T2',  '剂油比',   'feature_condition', False),
    # 原料油 (categorical)
    (91, 'oil', '原料油', 'feature_categorical', False),
    # 目标变量
    (58, 'C1',     '焦炭',   'target', False),
    (59, 'C2',     '干气',   'target', False),
    (60, 'C2-1',   '氢气',   'target', False),
    (61, 'C2-2',   '硫化氢', 'target', False),
    (62, 'C2-3',   '甲烷',   'target', False),
    (63, 'C2-4',   '乙烷',   'target', False),
    (64, 'C2-5',   '乙烯',   'target', False),
    (65, 'C3',     '液化气', 'target', False),
    (66, 'C3-1',   '丙烷',   'target', False),
    (67, 'C3-2',   '丙烯',   'target', False),
    (68, 'C3-3',   '正丁烷', 'target', False),
    (69, 'C3-4',   '异丁烷', 'target', False),
    (70, 'C3-5',   'C4烯烃', 'target', False),
    (71, 'C3-5-1', '1-丁烯', 'target', False),
    (72, 'C3-5-2', '异丁烯', 'target', False),
    (73, 'C3-5-3', '顺丁烯', 'target', False),
    (74, 'C3-5-4', '反丁烯', 'target', False),
    (75, 'C3-5-5', '丁二烯', 'target', False),
    (76, 'C4',     '汽油',   'target', False),
    (77, 'C5',     '柴油',   'target', False),
    (78, 'C6',     '重油',   'target', False),
    (79, 'C7',     '轻收',   'target', False),
    (80, 'C8',     '总液收', 'target', False),
    (81, 'C9',     '转化率', 'target', False),
]

wb_out = Workbook()

# ---- Sheet 1: ML_data ----
ws_ml = wb_out.active
ws_ml.title = 'ML_data'

header = [f'{code}_{name}' for _, code, name, _, _ in columns]
ws_ml.append(header)

written = 0
for row in data_rows:
    t1 = row[56] if 56 < len(row) else None
    t2 = row[57] if 57 < len(row) else None
    if t1 is None and t2 is None:
        continue
    out_row = []
    for col_idx, code, name, col_type, fill_zero in columns:
        val = row[col_idx] if col_idx < len(row) else None
        if val is None and fill_zero:
            val = 0
        out_row.append(val)
    ws_ml.append(out_row)
    written += 1

# ---- Sheet 2: Column_Info ----
ws_info = wb_out.create_sheet('Column_Info')
ws_info.append(['列名', '代码', '中文名', '类型说明', 'NaN处理', '原始列索引'])

type_desc = {
    'feature_composition': '特征-配方组成',
    'feature_property':    '特征-理化性质',
    'feature_condition':   '特征-操作条件',
    'feature_categorical': '特征-类别',
    'target':              '目标变量',
}
for col_idx, code, name, col_type, fill_zero in columns:
    col_name = f'{code}_{name}'
    fill_desc = '保留NaN（数据缺失）'
    ws_info.append([col_name, code, name, type_desc[col_type], fill_desc, col_idx])

wb_out.save('ml_catalyst_yield.xlsx')
print(f'Done. Written={written} rows, {len(columns)} columns')
print('File saved: d:/cc_test/ml_catalyst_yield.xlsx')
